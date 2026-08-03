import pandas as pd
import re
import sys
import os

def parse_error_message(msg):
    if not isinstance(msg, str) or msg.strip() == "":
        return "", "", "", "N/A"

    inv_code = device_type = device_error = ""
    table_frag = "N/A"

    inv_match = re.search(r'Inv Code\s*-\s*([^,]+)', msg, re.IGNORECASE)
    if inv_match:
        inv_code = inv_match.group(1).strip()

    type_match = re.search(r'Device Type\s*-\s*([^,]+)', msg, re.IGNORECASE)
    if type_match:
        device_type = type_match.group(1).strip()

    error_match = re.search(r'Device Error\s*-(.+?)(?:,Table Frag|$)', msg, re.IGNORECASE)
    if error_match:
        device_error = error_match.group(1).strip()

    frag_match = re.search(r'Table Frag\s*=\s*([^,]+)', msg, re.IGNORECASE)
    if frag_match:
        table_frag = frag_match.group(1).strip()

    return inv_code, device_type, device_error, table_frag

def main(input_file):
    df = pd.read_excel(input_file, sheet_name="Tickets for Analysis", header=2)
    df.dropna(how="all", inplace=True)

    if 'Error_Message' not in df.columns:
        print("ERROR: 'Error_Message' column not found in the Excel file.")
        print(f"Available columns: {list(df.columns)}")
        sys.exit(1)

    df['Error_Message'] = df['Error_Message'].fillna('')

    parsed = df['Error_Message'].apply(parse_error_message)
    df['Inv_Code_Parsed']    = parsed.apply(lambda x: x[0])
    df['Device_Type_Parsed'] = parsed.apply(lambda x: x[1])
    df['Device_Error']       = parsed.apply(lambda x: x[2])
    df['Table_Frag']         = parsed.apply(lambda x: x[3])

    base, ext = os.path.splitext(input_file)
    output_file = base + "_parsed" + ext

    df.to_excel(output_file, index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(output_file)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", start_color="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(output_file)
    print(f"Done! Output saved to: {output_file}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python parse_error_messages.py <path_to_excel_file.xlsx>")
        sys.exit(1)
    main(sys.argv[1])